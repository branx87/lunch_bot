# ##config.py
import os
import openpyxl
import pytz
import logging
import json
from datetime import datetime
from dotenv import load_dotenv

logger = logging.getLogger(__name__)

# Загружаем переменные из .env
load_dotenv()

# Общие настройки
TIMEZONE = pytz.timezone('Europe/Moscow')
CONFIG_FILE = "config.xlsx"
LOCATIONS = ["Офис", "ПЦ 1", "ПЦ 2", "Склад"]

def load_config():
    try:
        # Загружаем настройки из .env
        token = os.getenv('BOT_TOKEN')
        if not token:
            raise ValueError("Токен бота не указан в .env файле")
        
        admin_ids = [int(id_str) for id_str in os.getenv('ADMIN_IDS', '').split(',') if id_str]
        provider_ids = [int(id_str) for id_str in os.getenv('PROVIDER_IDS', '').split(',') if id_str]
        accounting_ids = [int(id_str) for id_str in os.getenv('ACCOUNTING_IDS', '').split(',') if id_str]
        
        # Остальные настройки загружаем из Excel
        if not os.path.exists(CONFIG_FILE):
            raise FileNotFoundError(f"Файл конфигурации {CONFIG_FILE} не найден")
        
        wb = openpyxl.load_workbook(CONFIG_FILE)
        ws = wb.active
        
        # Сотрудники (столбец G)
        staff_names = set()
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if row[6]:  # столбец G (индекс 6)
                name = ' '.join(str(row[6]).strip().split()).lower()
                staff_names.add(name)
                
                parts = name.split()
                if len(parts) >= 2:
                    reversed_name = f"{parts[1]} {parts[0]}"
                    if len(parts) > 2:
                        reversed_name += " " + " ".join(parts[2:])
                    staff_names.add(reversed_name)
        
        # Праздники (столбцы K и L)
        holidays = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[10] and row[11]:  # столбцы K (индекс 10) и L (индекс 11)
                try:
                    date_str = str(row[10]).strip()
                    try:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
                    except ValueError:
                        date_obj = datetime.strptime(date_str, "%d.%m.%Y").date()
                    holidays[date_obj.strftime("%Y-%m-%d")] = str(row[11]).strip()
                except Exception as e:
                    logger.warning(f"Не удалось обработать дату праздника: {row[10]}. Ошибка: {e}")
        
        # Меню (столбец I)
        menu = {}
        current_day = None
        dish_counter = 0
        
        # Начинаем с I2 (пропускаем заголовок)
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'I{row}'].value
            
            # Пропускаем пустые ячейки
            if cell_value is None:
                continue
                
            cell_value = str(cell_value).strip()
            
            # Определяем день недели
            if cell_value in ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]:
                current_day = cell_value
                menu[current_day] = {
                    "first": "",
                    "main": "",
                    "salad": ""
                }
                dish_counter = 0
                continue
                
            # Заполняем блюда для текущего дня
            if current_day:
                if dish_counter == 0:
                    menu[current_day]["first"] = cell_value
                elif dish_counter == 1:
                    menu[current_day]["main"] = cell_value
                elif dish_counter == 2:
                    menu[current_day]["salad"] = cell_value
                
                dish_counter += 1

        # Добавляем отсутствующие дни как None
        all_days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
        for day in all_days:
            if day not in menu:
                menu[day] = None

        logger.info(f"Загруженное меню:\n{json.dumps(menu, indent=2, ensure_ascii=False)}")
        
        return {
            'token': token,
            'admin_ids': admin_ids,
            'provider_ids': provider_ids,
            'accounting_ids': accounting_ids,
            'staff_names': staff_names,
            'holidays': holidays,
            'menu': menu
        }
    except Exception as e:
        logger.error(f"Критическая ошибка загрузки конфигурации: {e}", exc_info=True)
        raise

# Инициализация конфигурации ДОЛЖНА БЫТЬ ВНЕ функции load_config()
try:
    CONFIG = load_config()
    TOKEN = CONFIG['token']
    ADMIN_IDS = CONFIG['admin_ids']
    PROVIDER_IDS = CONFIG['provider_ids']
    ACCOUNTING_IDS = CONFIG['accounting_ids']
    STAFF_NAMES = CONFIG['staff_names']
    HOLIDAYS = CONFIG.get('holidays', {})
    MENU = CONFIG['menu']
except Exception as e:
    logger.error(f"Не удалось загрузить конфигурацию: {e}")
    exit(1)
