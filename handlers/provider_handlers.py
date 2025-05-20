# ##handlers/provider_handlers.py
import openpyxl
from telegram import ReplyKeyboardMarkup, Update
from telegram.ext import ConversationHandler, MessageHandler, filters
from telegram.ext import ContextTypes

from config import CONFIG, CONFIG_FILE
from constants import EDIT_MENU_DAY, EDIT_MENU_FIRST, EDIT_MENU_MAIN, EDIT_MENU_SALAD
from handlers.admin_config_handlers import cancel_config
from keyboards import create_provider_menu_keyboard

async def edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Инициирует процесс редактирования меню.
    Отображает клавиатуру с днями недели для выбора.
    Переводит в состояние EDIT_MENU_DAY.
    """
    await update.message.reply_text(
        "Выберите день недели:",
        reply_markup=ReplyKeyboardMarkup([
            ["Понедельник", "Вторник", "Среда"],
            ["Четверг", "Пятница", "Суббота"],
            ["Воскресенье", "❌ Отмена"]
        ], resize_keyboard=True)
    )
    return EDIT_MENU_DAY  # Строковая константа

async def handle_menu_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает выбор дня недели для редактирования.
    Проверяет корректность выбранного дня.
    Сохраняет выбранный день в user_data.
    Переводит в состояние EDIT_MENU_FIRST.
    """
    day = update.message.text
    if day not in ["Понедельник", "Вторник", ..., "Воскресенье"]:
        await update.message.reply_text("❌ Выберите день из списка")
        return EDIT_MENU_DAY
    
    context.user_data['edit_menu_day'] = day
    await update.message.reply_text("Введите первое блюдо:")
    return EDIT_MENU_FIRST

async def handle_menu_first(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает ввод первого блюда.
    Сохраняет значение в user_data.
    Переводит в состояние EDIT_MENU_MAIN.
    """
    context.user_data['first'] = update.message.text
    await update.message.reply_text("Введите основное блюдо:")
    return EDIT_MENU_MAIN

async def handle_menu_main(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Обрабатывает ввод основного блюда.
    Сохраняет значение в user_data.
    Переводит в состояние EDIT_MENU_SALAD.
    """
    context.user_data['main'] = update.message.text
    await update.message.reply_text("Введите салат:")
    return EDIT_MENU_SALAD

async def handle_menu_salad(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Завершает процесс редактирования меню.
    Сохраняет все изменения:
    - В конфигурации в памяти (CONFIG)
    - В Excel-файле (CONFIG_FILE)
    Возвращает пользователя в главное меню поставщика.
    """
    day = context.user_data['edit_menu_day']
    first = context.user_data['first']
    main = context.user_data['main']
    salad = update.message.text
    
    # Обновляем конфиг в памяти
    CONFIG['menu'][day] = {
        "first": first,
        "main": main,
        "salad": salad
    }
    
    # Обновляем Excel файл
    wb = openpyxl.load_workbook(CONFIG_FILE)
    ws = wb.active
    
    # Находим строку с нужным днем
    for row in range(2, ws.max_row + 1):
        if ws[f'I{row}'].value == day:
            ws[f'I{row+1}'] = first
            ws[f'I{row+2}'] = main
            ws[f'I{row+3}'] = salad
            break
    
    wb.save(CONFIG_FILE)
    
    await update.message.reply_text(
        f"✅ Меню на {day} обновлено!",
        reply_markup=create_provider_menu_keyboard()
    )
    return ConversationHandler.END

def setup_provider_handlers(application):
    """
    Настраивает и добавляет обработчики для функционала поставщиков:
    - ConversationHandler для редактирования меню
    - Проверяет права доступа (provider_ids)
    - Обрабатывает все этапы изменения меню
    - Поддерживает отмену операции
    """
    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(
            filters.Regex("^✏️ Изменить меню$") & 
            filters.User(user_id=CONFIG['provider_ids']),
            edit_menu
        )],
        states={
            EDIT_MENU_DAY: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu_day)],
            EDIT_MENU_FIRST: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu_first)],
            EDIT_MENU_MAIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu_main)],
            EDIT_MENU_SALAD: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_menu_salad)]
        },
        fallbacks=[
            MessageHandler(filters.Regex("^(❌ Отмена|Отмена)$"), cancel_config)
        ],
        allow_reentry=True
    )
    application.add_handler(conv_handler)