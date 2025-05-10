import logging
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ContextTypes, ConversationHandler, MessageHandler
from config import CONFIG, CONFIG_FILE, load_config
from keyboards import create_admin_keyboard, create_month_selection_keyboard
from admin import export_accounting_report
from .states import BROADCAST_MESSAGE, SELECT_MONTH_RANGE
from db import db
import asyncio
from filelock import FileLock
import openpyxl

logger = logging.getLogger(__name__)

async def handle_admin_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка выбора в админ-меню"""
    text = update.message.text.strip().lower()
    user = update.effective_user

    # Проверка прав администратора
    if user.id not in CONFIG.get('admin_ids', []) and user.id not in CONFIG.get('accounting_ids', []):
        await update.message.reply_text("❌ У вас нет прав для выполнения этой команды.")
        return ConversationHandler.END  # Заменили ADMIN_MESSAGE на завершение диалога

    if text == "📢 Сделать рассылку":
        await update.message.reply_text(
            "Введите сообщение для рассылки:",
            reply_markup=ReplyKeyboardMarkup([["Отмена"]], resize_keyboard=True)
        )
        return BROADCAST_MESSAGE

    elif text == "отмена":
        await update.message.reply_text(
            "Действие отменено",
            reply_markup=create_admin_keyboard()
        )
        return ConversationHandler.END  # Заменили ADMIN_MESSAGE

    elif text in ["📊 отчет за день", "отчет за день"]:
        if user.id in CONFIG.get('admin_ids', []) or user.id in CONFIG.get('accounting_ids', []):
            await export_accounting_report(update, context)
        else:
            await update.message.reply_text("❌ У вас нет прав для просмотра отчетов")
        return ConversationHandler.END  # Заменили ADMIN_MESSAGE

    elif text in ["📅 отчет за месяц", "отчет за месяц"]:
        if user.id in CONFIG.get('admin_ids', []) or user.id in CONFIG.get('accounting_ids', []):
            await update.message.reply_text(
                "Выберите период:",
                reply_markup=create_month_selection_keyboard()
            )
            return SELECT_MONTH_RANGE
        else:
            await update.message.reply_text("❌ У вас нет прав для просмотра отчетов")
        return ConversationHandler.END  # Заменили ADMIN_MESSAGE

    # Неизвестная команда
    await update.message.reply_text(
        "Неизвестная команда. Пожалуйста, используйте кнопки меню.",
        reply_markup=create_admin_keyboard()
    )
    return ConversationHandler.END  # Заменили ADMIN_MESSAGE

async def handle_broadcast_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка команды рассылки"""
    user = update.effective_user
    if user.id not in CONFIG['admin_ids']:
        await update.message.reply_text("❌ У вас нет прав для этой команды")
        return
    
    await update.message.reply_text(
        "Введите сообщение для рассылки:",
        reply_markup=ReplyKeyboardMarkup([["❌ Отмена"]], resize_keyboard=True)
    )
    return "BROADCAST_MESSAGE"

async def process_broadcast_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка текста рассылки"""
    text = update.message.text
    
    if text.lower() in ["отменить рассылку", "❌ отменить рассылку"]:
        await update.message.reply_text(
            "❌ Рассылка отменена",
            reply_markup=create_admin_keyboard()
        )
        return ConversationHandler.END
    
    try:
        # Получаем всех верифицированных пользователей
        db.cursor.execute("SELECT telegram_id, full_name FROM users WHERE is_verified = TRUE")
        users = db.cursor.fetchall()
        
        if not users:
            await update.message.reply_text(
                "❌ Нет пользователей для рассылки",
                reply_markup=create_admin_keyboard()
            )
            return ConversationHandler.END
        
        # Отправляем уведомление о начале рассылки
        await update.message.reply_text(f"⏳ Начинаю рассылку для {len(users)} пользователей...")
        
        success = 0
        failed = []
        
        for user_id, full_name in users:
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=f"📢 Сообщение от администратора:\n\n{text}"
                )
                success += 1
                await asyncio.sleep(0.1)  # Задержка между сообщениями
            except Exception as e:
                failed.append(f"{full_name} (ID: {user_id})")
                logger.error(f"Ошибка отправки {user_id}: {e}")
        
        # Формируем отчет
        report = (
            f"✅ Рассылка завершена:\n"
            f"• Всего получателей: {len(users)}\n"
            f"• Успешно: {success}\n"
            f"• Не удалось: {len(failed)}"
        )
        
        if failed:
            report += "\n\nНе удалось отправить:\n" + "\n".join(failed[:10])  # Показываем первые 10 ошибок
        
        await update.message.reply_text(
            report,
            reply_markup=create_admin_keyboard()
        )
        
    except Exception as e:
        logger.error(f"Ошибка рассылки: {e}")
        await update.message.reply_text(
            "❌ Произошла ошибка при рассылке",
            reply_markup=create_admin_keyboard()
        )
    
    return ConversationHandler.END
    

async def add_user_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Добавляет ID пользователя в конфигурационный файл
    Формат команды: /add_user <admin/provider/accounting> <user_id>
    """
    user = update.effective_user
    
    # Проверка прав
    if user.id not in CONFIG['admin_ids']:
        await update.message.reply_text("❌ Только администраторы могут добавлять пользователей")
        return

    try:
        # Парсим аргументы
        args = context.args
        if len(args) != 2:
            raise ValueError(
                "Неверный формат команды.\n"
                "Используйте: /add_user <тип> <id>\n"
                "Типы: admin, provider, accounting"
            )

        user_type = args[0].lower()
        new_id = int(args[1])  # Проверка что ID - число

        # Определяем столбец в Excel
        column_map = {
            'admin': 'B',
            'provider': 'C',
            'accounting': 'D'
        }

        if user_type not in column_map:
            raise ValueError("Неверный тип пользователя. Допустимо: admin, provider, accounting")

        column = column_map[user_type]

        # Блокировка файла для безопасной записи
        with FileLock(CONFIG_FILE + ".lock"):
            # Открываем Excel-файл
            wb = openpyxl.load_workbook(CONFIG_FILE)
            ws = wb.active
            
            # Ищем первую пустую ячейку в столбце
            row = 2
            while ws[f'{column}{row}'].value is not None:
                row += 1

            # Записываем новый ID
            ws[f'{column}{row}'] = new_id
            wb.save(CONFIG_FILE)

        # Обновляем конфиг в памяти
        global CONFIG
        CONFIG = load_config()
        
        # Обновляем соответствующий список ID
        if user_type == 'admin':
            CONFIG['admin_ids'].append(new_id)
        elif user_type == 'provider':
            CONFIG['provider_ids'].append(new_id)
        elif user_type == 'accounting':
            CONFIG['accounting_ids'].append(new_id)

        await update.message.reply_text(
            f"✅ Пользователь {new_id} успешно добавлен как {user_type}!\n"
            f"Текущие {user_type} IDs: {CONFIG[user_type + '_ids']}"
        )

    except ValueError as ve:
        await update.message.reply_text(f"❌ Ошибка ввода: {str(ve)}")
    except Exception as e:
        logger.error(f"Ошибка добавления пользователя: {e}", exc_info=True)
        await update.message.reply_text("❌ Произошла ошибка при добавлении пользователя")
        
__all__ = [
    'handle_admin_choice',
    'handle_broadcast_command',
    'process_broadcast_message',
    'add_user_id'  # Добавляем новую функцию в экспорт
]
