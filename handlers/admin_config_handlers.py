# ##handlers/admin_config_handlers.py
from asyncio.log import logger
from datetime import datetime
import os
import re
import openpyxl  # Добавляем импорт
from config import CONFIG, CONFIG_FILE
from constants import ADD_ACCOUNTANT, ADD_ADMIN, ADD_HOLIDAY_DATE, ADD_HOLIDAY_NAME, ADD_PROVIDER, ADD_STAFF, CONFIG_MENU
from dotenv import load_dotenv
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ConversationHandler, MessageHandler, filters
from telegram.ext import ContextTypes

from keyboards import create_admin_config_keyboard, create_admin_keyboard

# Инициализация dotenv
load_dotenv()

def update_env_file(key: str, value: str):
    """Обновляет значение переменной в .env файле. Если переменной нет, она будет добавлена."""
    env_file = ".env"
    if not os.path.exists(env_file):
        with open(env_file, 'w'): pass
    
    lines = []
    updated = False
    
    if os.path.exists(env_file):
        with open(env_file, 'r') as file:
            lines = file.readlines()
    
    with open(env_file, 'w') as file:
        for line in lines:
            if line.startswith(f"{key}="):
                file.write(f"{key}={value}\n")
                updated = True
            else:
                file.write(line)
        if not updated:
            file.write(f"{key}={value}\n")

async def config_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отображает главное меню управления конфигурацией и проверяет права пользователя."""
    if update.effective_user.id not in CONFIG['admin_ids']:
        await update.message.reply_text("❌ У вас нет прав")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "⚙️ Управление конфигурацией:",
        reply_markup=create_admin_config_keyboard()
    )
    return CONFIG_MENU

async def start_add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запрашивает у пользователя Telegram ID для добавления нового администратора."""
    await update.message.reply_text("Введите Telegram ID нового администратора:")
    return ADD_ADMIN

async def handle_add_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенный Telegram ID администратора, проверяет его валидность и добавляет в .env файл."""
    try:
        new_id = update.message.text.strip()
        if not re.match(r'^\d+$', new_id):
            raise ValueError
        
        current_ids = os.getenv('ADMIN_IDS', '').split(',')
        current_ids = [id.strip() for id in current_ids if id.strip()]
        
        if new_id in current_ids:
            await update.message.reply_text("⚠️ Этот ID уже есть в списке")
            return ADD_ADMIN
        
        current_ids.append(new_id)
        new_ids_str = ','.join(current_ids)
        
        # Обновляем .env файл
        update_env_file('ADMIN_IDS', new_ids_str)
        
        # Обновляем конфиг в памяти
        CONFIG['admin_ids'] = [int(id) for id in current_ids]
        
        await update.message.reply_text(
            f"✅ ID {new_id} добавлен в администраторы",
            reply_markup=create_admin_config_keyboard()
        )
        return CONFIG_MENU
    except ValueError:
        await update.message.reply_text("❌ Введите корректный Telegram ID (только цифры)")
        return ADD_ADMIN

# Аналогичные функции для поставщиков и бухгалтеров
async def start_add_provider(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запрашивает у пользователя Telegram ID для добавления нового поставщика."""
    await update.message.reply_text("Введите Telegram ID нового поставщика:")
    return ADD_PROVIDER

async def handle_add_provider(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенный Telegram ID поставщика, проверяет его валидность и добавляет в .env файл."""
    try:
        new_id = update.message.text.strip()
        if not re.match(r'^\d+$', new_id):
            raise ValueError
        
        current_ids = os.getenv('PROVIDER_IDS', '').split(',')
        current_ids = [id.strip() for id in current_ids if id.strip()]
        
        if new_id in current_ids:
            await update.message.reply_text("⚠️ Этот ID уже есть в списке")
            return ADD_PROVIDER
        
        current_ids.append(new_id)
        new_ids_str = ','.join(current_ids)
        
        update_env_file('PROVIDER_IDS', new_ids_str)
        CONFIG['provider_ids'] = [int(id) for id in current_ids]
        
        await update.message.reply_text(
            f"✅ ID {new_id} добавлен в поставщики",
            reply_markup=create_admin_config_keyboard()
        )
        return CONFIG_MENU
    except ValueError:
        await update.message.reply_text("❌ Введите корректный Telegram ID (только цифры)")
        return ADD_PROVIDER

async def start_add_accountant(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запрашивает у пользователя Telegram ID для добавления нового бухгалтера."""
    await update.message.reply_text("Введите Telegram ID нового бухгалтера:")
    return ADD_ACCOUNTANT

async def handle_add_accountant(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенный Telegram ID бухгалтера, проверяет его валидность и добавляет в .env файл."""
    try:
        new_id = update.message.text.strip()
        if not re.match(r'^\d+$', new_id):
            raise ValueError
        
        current_ids = os.getenv('ACCOUNTING_IDS', '').split(',')
        current_ids = [id.strip() for id in current_ids if id.strip()]
        
        if new_id in current_ids:
            await update.message.reply_text("⚠️ Этот ID уже есть в списке")
            return ADD_ACCOUNTANT
        
        current_ids.append(new_id)
        new_ids_str = ','.join(current_ids)
        
        update_env_file('ACCOUNTING_IDS', new_ids_str)
        CONFIG['accounting_ids'] = [int(id) for id in current_ids]
        
        await update.message.reply_text(
            f"✅ ID {new_id} добавлен в бухгалтерию",
            reply_markup=create_admin_config_keyboard()
        )
        return CONFIG_MENU
    except ValueError:
        await update.message.reply_text("❌ Введите корректный Telegram ID (только цифры)")
        return ADD_ACCOUNTANT

async def start_add_staff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запрашивает у пользователя ФИ нового сотрудника для добавления в базу."""
    await update.message.reply_text(
        "Введите ФИ нового сотрудника (например, Иванов Иван):",
        reply_markup=ReplyKeyboardMarkup([["❌ Отмена"]], resize_keyboard=True)
    )
    return ADD_STAFF

async def handle_add_staff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенное ФИО сотрудника, проверяет его уникальность и добавляет в Excel-файл."""
    try:
        full_name = ' '.join(update.message.text.strip().split())
        if not full_name or len(full_name.split()) < 2:
            await update.message.reply_text("❌ Введите полное ФИО (минимум имя и фамилию)")
            return ADD_STAFF
        
        # Проверяем, существует ли уже такой сотрудник
        wb = openpyxl.load_workbook(CONFIG_FILE)
        ws = wb.active
        
        # Проверяем все записи в столбце G
        existing_staff = set()
        for row in range(2, ws.max_row + 1):
            if ws[f'G{row}'].value:
                existing_staff.add(ws[f'G{row}'].value.strip().lower())
        
        # Проверяем все варианты имени
        name_variants = {
            full_name.lower(),
            ' '.join(full_name.split()[::-1]).lower()  # Перевернутое ФИО
        }
        
        # Проверяем на совпадение
        if any(variant in existing_staff for variant in name_variants):
            await update.message.reply_text(
                "❌ Такой сотрудник уже существует",
                reply_markup=create_admin_config_keyboard()
            )
            return CONFIG_MENU
        
        # Находим первую пустую строку в столбце G
        row = 2
        while ws[f'G{row}'].value is not None:
            row += 1
        
        # Добавляем сотрудника
        ws[f'G{row}'] = full_name
        wb.save(CONFIG_FILE)
        
        # Обновляем конфиг в памяти
        CONFIG['staff_names'].update(name_variants)
        
        await update.message.reply_text(
            f"✅ Сотрудник '{full_name}' добавлен",
            reply_markup=create_admin_config_keyboard()
        )
        return CONFIG_MENU
        
    except Exception as e:
        logger.error(f"Ошибка при добавлении сотрудника: {e}")
        await update.message.reply_text(
            "❌ Произошла ошибка. Попробуйте позже.",
            reply_markup=create_admin_config_keyboard()
        )
        return CONFIG_MENU

async def start_add_holiday(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Запрашивает у пользователя дату праздника для добавления в базу."""
    await update.message.reply_text(
        "Введите дату праздника в формате ДД.ММ.ГГГГ:",
        reply_markup=ReplyKeyboardMarkup([["❌ Отмена"]], resize_keyboard=True)
    )
    return ADD_HOLIDAY_DATE

async def handle_holiday_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенную дату праздника, проверяет ее формат и сохраняет для дальнейшего использования."""
    try:
        date_str = update.message.text.strip()
        date_obj = datetime.strptime(date_str, "%d.%m.%Y").date()
        context.user_data['holiday_date'] = date_obj.strftime("%Y-%m-%d")
        
        await update.message.reply_text(
            "Введите название праздника:",
            reply_markup=ReplyKeyboardMarkup([["❌ Отмена"]], resize_keyboard=True)
        )
        return ADD_HOLIDAY_NAME
    except ValueError:
        await update.message.reply_text("❌ Неверный формат даты. Используйте ДД.ММ.ГГГГ")
        return ADD_HOLIDAY_DATE

async def handle_holiday_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает введенное название праздника и сохраняет его вместе с датой в Excel-файл."""
    holiday_name = update.message.text.strip()
    if not holiday_name:
        await update.message.reply_text("❌ Название праздника не может быть пустым")
        return ADD_HOLIDAY_NAME
    
    date_str = context.user_data['holiday_date']
    
    # Добавляем в конфиг (в памяти)
    CONFIG['holidays'][date_str] = holiday_name
    
    # Добавляем в Excel (столбцы K и L)
    wb = openpyxl.load_workbook(CONFIG_FILE)
    ws = wb.active
    
    # Находим первую пустую строку в столбцах K и L
    row = 2
    while ws[f'K{row}'].value is not None:
        row += 1
    
    ws[f'K{row}'] = date_str
    ws[f'L{row}'] = holiday_name
    wb.save(CONFIG_FILE)
    
    await update.message.reply_text(
        f"✅ Праздник '{holiday_name}' на {date_str} добавлен",
        reply_markup=create_admin_config_keyboard()
    )
    return CONFIG_MENU

async def cancel_config(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отменяет текущую операцию изменения конфигурации и возвращает пользователя в главное меню."""
    await update.message.reply_text(
        "✅ Изменения сохранены",
        reply_markup=create_admin_keyboard()
    )
    return ConversationHandler.END

def setup_admin_config_handlers(application):
    """Настраивает и добавляет обработчики команд для управления конфигурацией в приложение."""
    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(
            filters.Regex("^⚙️ Управление конфигурацией$") & 
            filters.User(user_id=CONFIG['admin_ids']),
            config_menu
        )],
        states={
            CONFIG_MENU: [
                MessageHandler(filters.Regex("^➕ Добавить администратора$"), start_add_admin),
                MessageHandler(filters.Regex("^➕ Добавить поставщика$"), start_add_provider),
                MessageHandler(filters.Regex("^➕ Добавить бухгалтера$"), start_add_accountant),
                MessageHandler(filters.Regex("^➕ Добавить сотрудника$"), start_add_staff),
                MessageHandler(filters.Regex("^➕ Добавить праздник$"), start_add_holiday),
            ],
            ADD_ADMIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_add_admin)],
            ADD_PROVIDER: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_add_provider)],
            ADD_ACCOUNTANT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_add_accountant)],
            ADD_STAFF: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_add_staff)],
            ADD_HOLIDAY_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_holiday_date)],
            ADD_HOLIDAY_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_holiday_name)],
        },
        fallbacks=[
            MessageHandler(filters.Regex("^(❌ Отмена|Отмена)$"), cancel_config),
            MessageHandler(filters.Regex("^🏠 Главное меню$"), cancel_config)
        ],
        allow_reentry=True
    )
    application.add_handler(conv_handler)