import aiocron
from config import TIMEZONE, CONFIG
from datetime import datetime, timedelta
import logging
from db import db
from admin import export_accounting_report
from telegram.ext import ContextTypes
import openpyxl

logger = logging.getLogger(__name__)

async def send_reminder_to_users(application):
    try:
        now = datetime.now(TIMEZONE)
        
        if now.weekday() < 5:  # Пн-Пт
            logger.info(f"Запуск напоминаний в {now}")
            
            db.cursor.execute("""
                SELECT telegram_id 
                FROM users 
                WHERE is_verified = TRUE 
                AND telegram_id NOT IN (
                    SELECT u.telegram_id 
                    FROM users u 
                    JOIN orders o ON u.id = o.user_id 
                    WHERE o.order_date = ? AND o.is_preliminary = FALSE
                )
            """, (now.date().isoformat(),))
            
            users = db.cursor.fetchall()
            logger.info(f"Найдено {len(users)} пользователей без заказов")
            
            for user in users:
                user_id = user[0]
                try:
                    await application.bot.send_message(
                        chat_id=user_id,
                        text="⏰ *Не забудьте заказать обед!* 🍽\n\n"
                             "Прием заказов открыт до 9:30.\n\n"
                             "Заказы принимаются через *БИТРИКС*, бот в тестовом режиме!",
                        parse_mode="Markdown"
                    )
                except Exception as e:
                    logger.error(f"Ошибка при отправке напоминания пользователю {user_id}: {e}")
                    
    except Exception as e:
        logger.error(f"Ошибка в send_reminder_to_users: {e}")

async def send_provider_report(application):
    """Отправка ежедневного отчёта поставщикам (в 9:30 по будням)"""
    try:
        logger.info("Запуск отправки отчета поставщикам")

        now = datetime.now(TIMEZONE)
        today = now.date().isoformat()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Заказы на сегодня"
        ws.append(["Объект", "Количество порций"])

        # SQL-запрос: уникальные локации с заказами за сегодня
        db.cursor.execute('''
            SELECT u.location, SUM(o.quantity)
            FROM orders o
            JOIN users u ON o.user_id = u.id
            WHERE o.order_date = ?
              AND o.is_cancelled = FALSE
            GROUP BY u.location
            ORDER BY u.location
        ''', (today,))

        locations_with_orders = []
        total = 0

        for row in db.cursor.fetchall():
            ws.append([row[0], row[1]])
            locations_with_orders.append(row[0])
            total += row[1]

        unique_locations_count = len(locations_with_orders)

        # Автоподбор ширины столбцов
        for col in ws.columns:
            max_length = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = (max_length + 2) * 1.2

        # Сохраняем файл
        file_path = f"provider_report_{today.replace('-', '')}.xlsx"
        wb.save(file_path)

        # Рассылаем всем поставщикам
        success = 0
        with open(file_path, 'rb') as file:
            for provider_id in CONFIG.get('provider_ids', []):
                try:
                    await application.bot.send_document(
                        chat_id=provider_id,
                        document=file,
                        caption=(
                            f"🍽 Заказы на {now.strftime('%d.%m.%Y')}\n"
                            f"📍 Локаций: {unique_locations_count} | 🍛 Всего: {total} порций\n"
                            "📋 Детализация в приложенном файле"
                        )
                    )
                    success += 1
                    file.seek(0)  # Сброс позиции файла для следующего получателя
                except Exception as e:
                    logger.error(f"Ошибка отправки поставщику {provider_id}: {e}")

        logger.info(f"Отправлено {success}/{len(CONFIG.get('provider_ids', []))} поставщикам")

    except Exception as e:
        logger.error(f"Ошибка отправки отчета поставщикам: {e}")

async def send_accounting_report(application):
    """Отправка ежемесячного отчёта бухгалтерии (в 11:00 последнего дня месяца)"""
    try:
        logger.info("Запуск отправки отчета бухгалтерии")

        now = datetime.now(TIMEZONE)
        today = now.date()

        # Определяем начало и конец месяца
        if today.day == 1:
            # Сегодня первое число — отправляем за предыдущий месяц
            first_day_current_month = now.replace(day=1)
            last_day_prev_month = first_day_current_month - timedelta(days=1)
            start_date = last_day_prev_month.replace(day=1).date()
            end_date = last_day_prev_month.date()
        else:
            # Иначе отправляем за текущий месяц до сегодняшнего числа
            start_date = now.replace(day=1).date()
            end_date = today

        wb = openpyxl.Workbook()

        # 1. Лист "Детализация"
        ws_detailed = wb.active
        ws_detailed.title = "Детализация"
        detailed_headers = ["ФИО", "Объект", "Дата", "Количество", "Тип заказа"]
        ws_detailed.append(detailed_headers)
        ws_detailed.auto_filter.ref = "A1:E1"

        total_portions = 0

        db.cursor.execute('''
            SELECT 
                u.full_name,
                u.location,
                o.order_date,
                o.quantity,
                CASE WHEN o.is_preliminary THEN 'Предзаказ' ELSE 'Обычный' END
            FROM orders o
            JOIN users u ON o.user_id = u.id
            WHERE o.order_date BETWEEN ? AND ?
              AND o.is_cancelled = FALSE
            ORDER BY o.order_date, u.full_name
        ''', (start_date.isoformat(), end_date.isoformat()))

        for row in db.cursor.fetchall():
            formatted_date = datetime.strptime(row[2], "%Y-%m-%d").strftime("%d.%m.%Y")
            ws_detailed.append([
                row[0],  # ФИО
                row[1],  # Объект
                formatted_date,  # Дата
                row[3],  # Количество
                row[4]   # Тип
            ])
            total_portions += row[3]

        # 2. Лист "Сводка по сотрудникам"
        ws_summary_users = wb.create_sheet("Сводка по сотрудникам")
        summary_users_headers = ["ФИО", "Объект", "Всего порций"]
        ws_summary_users.append(summary_users_headers)
        ws_summary_users.auto_filter.ref = "A1:C1"

        db.cursor.execute('''
            SELECT 
                u.full_name,
                u.location,
                SUM(o.quantity)
            FROM orders o
            JOIN users u ON o.user_id = u.id
            WHERE o.order_date BETWEEN ? AND ?
              AND o.is_cancelled = FALSE
            GROUP BY u.full_name, u.location
            ORDER BY SUM(o.quantity) DESC
        ''', (start_date.isoformat(), end_date.isoformat()))

        for row in db.cursor.fetchall():
            ws_summary_users.append(row)

        # 3. Лист "Сводка по объектам"
        ws_summary_locations = wb.create_sheet("Сводка по объектам")
        ws_summary_locations.append(["Объект", "Порции"])
        ws_summary_locations.auto_filter.ref = "A1:B1"

        db.cursor.execute('''
            SELECT u.location, SUM(o.quantity)
            FROM orders o
            JOIN users u ON o.user_id = u.id
            WHERE o.order_date BETWEEN ? AND ?
              AND o.is_cancelled = FALSE
            GROUP BY u.location
            ORDER BY SUM(o.quantity) DESC
        ''', (start_date.isoformat(), end_date.isoformat()))

        for location, portions in db.cursor.fetchall():
            ws_summary_locations.append([location, portions])

        ws_summary_locations.append(["ВСЕГО", total_portions])

        # 4. Лист "Итоги"
        ws_stats = wb.create_sheet("Итоги")
        stats_headers = ["Показатель", "Значение"]
        ws_stats.append(stats_headers)

        stats_data = [
            ["Период", f"{start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"],
            ["Всего порций", total_portions],
            ["Уникальных сотрудников", ws_summary_users.max_row - 1],
            ["Дата формирования", datetime.now(TIMEZONE).strftime("%d.%m.%Y %H:%M")]
        ]
        for row in stats_data:
            ws_stats.append(row)

        # Форматирование
        from openpyxl.styles import Font
        bold_font = Font(bold=True)

        for sheet in wb.worksheets:
            # Жирные заголовки
            for row in sheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = bold_font

            # Автоподбор ширины
            for column in sheet.columns:
                max_length = max((len(str(cell.value)) if cell.value else 0 for cell in column), default=0)
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Сохраняем файл
        file_path = f"accounting_report_{start_date.strftime('%Y%m%d')}"
        if start_date != end_date:
            file_path += f"_to_{end_date.strftime('%Y%m%d')}"
        file_path += ".xlsx"
        wb.save(file_path)

        # Отправляем каждому бухгалтеру
        success = 0
        with open(file_path, 'rb') as file:
            for accounting_id in CONFIG.get('accounting_ids', []):
                try:
                    await application.bot.send_document(
                        chat_id=accounting_id,
                        document=file,
                        caption=(
                            f"📊 Бухгалтерский отчет\n"
                            f"📅 Период: {start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}\n"
                            f"🍽 Всего порций: {total_portions}"
                        )
                    )
                    success += 1
                    file.seek(0)  # Сбрасываем позицию файла для следующего получателя
                except Exception as e:
                    logger.error(f"Ошибка отправки бухгалтеру {accounting_id}: {e}")

        logger.info(f"Отправлено {success}/{len(CONFIG.get('accounting_ids', []))} бухгалтерам")

    except Exception as e:
        logger.error(f"Ошибка при отправке отчета бухгалтерии: {e}", exc_info=True)

async def setup_cron_jobs(application):
    # Напоминание пользователям в 9:00 по будням
    # @aiocron.crontab('* * * * *', tz=TIMEZONE)
    @aiocron.crontab('0 9 * * 1-5', tz=TIMEZONE)
    async def morning_reminder():
        await send_reminder_to_users(application)

    # Отчет поставщикам в 9:30 по рабочим дням
    # @aiocron.crontab('* * * * *', tz=TIMEZONE)
    @aiocron.crontab('30 9 * * 1-5', tz=TIMEZONE)
    async def provider_report():
        await send_provider_report(application)
    
    # Отчет бухгалтерии в 11:00 последнего дня месяца
    # @aiocron.crontab('* * * * *', tz=TIMEZONE)
    @aiocron.crontab('0 11 L * *', tz=TIMEZONE)
    async def accounting_report():
        await send_accounting_report(application)
